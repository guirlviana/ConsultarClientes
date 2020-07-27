import openpyxl
def consulta():
    planilha = openpyxl.load_workbook('Clientes 2.xlsx')
    planilha_clientes = planilha['Clientes']    
    allid = list()
    for linha in planilha_clientes.iter_rows(min_row= 11, min_col=3, max_col=3):    
        for cedula in linha:
            allid.append(cedula.value)
    allid.remove(None)
    cliente = int(input("Qual ID cliente: "))
    if cliente not in allid:
      print("Cliente nao encontrado")
      allid.clear()
      consulta()
    else:
        verDados(cliente, planilha_clientes)

def verDados(cliente, planilha_clientes):
    c = 11
    for linha in planilha_clientes.iter_rows(min_row= 11, min_col=3, max_col=3):    
        for cedula in linha:        
            if cedula.value == cliente:   
                print('=-' * 20)         
                print(f"I͇D͇: {cedula.value}")
                print(f"N͇O͇M͇E͇: {planilha_clientes[f'D{c}'].value}")
                print(f"T͇E͇L͇: {planilha_clientes[f'E{c}'].value}")
                print(f"E͇N͇D͇: {planilha_clientes[f'F{c}'].value}")
                print('=-' * 20) 
            else:
                c += 1

def pegarDados():
    while True:
        nome = str(input('NOME: '))
        if nome not in '':
            break
    while True:    
        tel = str(input('TEL: '))
        if tel not in '':
            break
    while True:    
        end = str(input('END: ')) 
        if end not in '':
            break
    cadastar(nome, tel, end)




def cadastar(nome, tel, end):
    planilha = openpyxl.load_workbook('Clientes 2.xlsx')
    planilha_clientes = planilha['Clientes']   
    allid = list()
    for linha in planilha_clientes.iter_rows(min_row= 11, min_col=3, max_col=3):    
        for cedula in linha:
            if cedula.value == None:
                pass
            else:
                allid.append(cedula.value)
    
    allid.sort()
    idusuario = allid[-1] +1
    linha_dado = len(allid) +11
    print(f'id novousuario = {idusuario}')
    print(f"linha a ser escrita = {linha_dado}")
    planilha_clientes[f'C{linha_dado}'] = idusuario
    planilha_clientes[f'D{linha_dado}'] = nome
    planilha_clientes[f'E{linha_dado}'] = tel
    planilha_clientes[f'F{linha_dado}'] = end
    planilha.save('Clientes 2.xlsx')




if __name__ == "__main__":
    print('=-' * 14)
    print('   CLIENTES HOMEBURGUER')
    print('=-' * 14)
    
    while True:
        escolha = str(input('[1] BUSCAR CLIENTE\n[2] CADASTRAR CLIENTE\nESCOLHA: '))
        if escolha in '12':
            break

    if escolha == '1':
        consulta()
    if escolha == '2':
        pegarDados()