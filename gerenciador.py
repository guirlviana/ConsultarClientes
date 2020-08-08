import openpyxl

class Manager():
    def __init__(self):
        self.planilha = openpyxl.load_workbook('Clientes test.xlsx')
        self.planilha_clientes = self.planilha['Clientes'] 

    def consulta(self):
        allid = list()
        for linha in self.planilha_clientes.iter_rows(min_row= 11, min_col=3, max_col=3):    
            for cedula in linha:
                allid.append(cedula.value)
        allid.remove(None)
        cliente = int(input("Qual ID cliente: "))
        if cliente not in allid:
            print("Cliente nao encontrado")
            allid.clear()
            self.consulta()
        else:
            self.verDados(cliente)

    def verDados(self, cliente):
        c = 11
        for linha in self.planilha_clientes.iter_rows(min_row= 11, min_col=3, max_col=3):    
            for cedula in linha:        
                if cedula.value == cliente: 
                    print('=-' * 20)         
                    print(f"ID: {cedula.value}")
                    print(f"NOME: {self.planilha_clientes[f'D{c}'].value}")                 
                    print(f"TEL: {self.planilha_clientes[f'E{c}'].value}")
                    print(f"END: {self.planilha_clientes[f'F{c}'].value}")
                    print('=-' * 20) 
                else:
                    c += 1

    def pegarDados(self):
        while True:
            nome = str(input('NOME: '))
            if nome not in '':
                break
        while True:    
            tel = str(input('TEL: '))
            if tel not in '':
                break
        while True:    
            end = str(input('END: ')) 
            if end not in '':
                break
        self.cadastar(nome, tel, end)


    def cadastar(self, nome, tel, end):
           
        allid = list()
        for linha in self.planilha_clientes.iter_rows(min_row= 11, min_col=3, max_col=3):    
            for cedula in linha:
                if cedula.value == None:
                    pass
                else:
                    allid.append(cedula.value)
        
        allid.sort()
        idusuario = allid[-1] +1
        linha_dado = len(allid) +11  
        print('=-' * 20)
        print(f'Adicionar id no contato = {idusuario}')
        print('=-' * 20)
        self.planilha_clientes[f'C{linha_dado}'] = idusuario
        self.planilha_clientes[f'D{linha_dado}'] = nome
        self.planilha_clientes[f'E{linha_dado}'] = tel
        self.planilha_clientes[f'F{linha_dado}'] = end
        self.planilha.save('Clientes test.xlsx')


    def menu(self):
        import os
        print('   CLIENTES HOMEBURGUER')
        print('=-' * 20)
        
        while True:
            escolha = str(input('[1] BUSCAR CLIENTE\n[2] CADASTRAR CLIENTE\n[3] LIMPAR TELA\n[4] SAIR\nESCOLHA: '))
            if escolha in '1234':
                break
        if escolha == '1':
            self.consulta()        
            self.menu()
        if escolha == '2':
            self.pegarDados()        
            self.menu()
        if escolha == '3':
            os.system('cls')
            self.menu()
        if escolha == '4':
            while True:
                decisão = str(input('Deseja sair? [S/N] ')).strip().upper()[0]
                if decisão in 'SN':
                    break
            if decisão == 'S':  
                os.system('exit')
            elif decisão == 'N':
                print()
                self.menu()    

if __name__ == "__main__":
    system = Manager()
    system.menu()